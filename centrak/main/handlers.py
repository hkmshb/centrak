import io
import os
import sys
import openpyxl
from datetime import datetime
from django.conf import settings
from django.db import connections, transaction
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from ezaddress.models import State

from sqlalchemy import create_engine
from dolfin.core import Storage
from dolfin.data import XlSheet
from core.models import UserProfile, BusinessOffice, Station, Powerline, \
        Voltage

User = get_user_model()

# python version flag
py3k = sys.version_info.major > 2

ERROR_NOT_INT = 'is invalid integer'
ERROR_NOT_DATE = 'has no date in it'
ERROR_NOT_FLOAT = 'is not a float'
ERROR_NOT_BOOLEAN = 'is not a boolean'
ERROR_MISSING_REQUIRED_TEXT = 'missing required text'
ERROR_NOT_UNICODE_OR_ASCII = 'not unicode or ascii string'
ERROR_NOT_ENUM_OF = lambda e: 'is bad integer/text for enum `%s`' % e.__name__

no_data = object()
no_enum = object()

def to_text(value):
    if isinstance(value, float):
        if int(value) == value:
            value = int(value)
    try:
        value = str(value)
    except UnicodeError:
        value = None
    return value


def to_bool(value):
    if value is not None:
        if type(value) == type(True):
            return value
        elif type(value) == type(0):
            return bool(value)

        # hack: needed for py2
        elif not py3k and 'long' in str(type(value)):
            return bool(value)
        
        value = to_text(value)
        if not value:
            return None
        if value.upper() in ['TRUE', 'YES', 'T', 'Y']:
            return True
        elif value.upper() in ['FALSE', 'NO', 'F', 'N']:
            return False
    return None


def to_enum(enum_type, value):
    if value is not None:
        if type(value) == type(0) or (not py3k and 'long' in str(type(value))):
            try:
                return enum_type(value)
            except ValueError:
                return None
        value = to_text(value)
        if value and hasattr(enum_type, value):
            return enum_type[value]
    return None


def handle_uploaded_files(upl_obj):
    dtfolder = datetime.today().strftime("%Y%m%d")
    dir_root = os.path.join(settings.MEDIA_ROOT, dtfolder)
    if not os.path.exists(dir_root):
        os.makedirs(dir_root)
    
    dtname = datetime.today().strftime("%H%M%S")
    filename = "%s-%s" % (dtname, upl_obj.name)

    filepath = os.path.join(dir_root, filename)
    try:
        with open(filepath, 'wb+') as destination:
            for chunk in upl_obj.chunks():
                destination.write(chunk)
            destination.flush()
        return filepath
    except:
        return None


class IXlrdHandlerBase(object):
    sheet_name = None
    table_name = None

    def __init__(self):
        self.errors = []
    
    def error(self, message):
        self.errors.append((self.sheet_name, message))

    def get_db_connstr(self):
        db_settings = settings.DATABASES['default']
        if 'postgresql' not in db_settings['ENGINE']:
            raise Exception("Backend database should be 'PostgreSQL'.")
        
        conn_str = "postgresql://%(USER)s:%(PASSWORD)s@%(HOST)s/%(NAME)s"
        return conn_str % db_settings
    
    def validate_column_headings(self):
        headers = (self.headers or ('fake-@#$%^',))
        xlsheet = XlSheet(self._filepath, self.sheet_name)
        if not XlSheet.find_headers(xlsheet, headers):
            raise ValidationError("Expected columns not found.")


class IPyXlHandlerBase(object):
    sheet_name = None

    def __init__(self, context, progress_callback=None):
        assert 'db' in context
        assert 'cache' in context
        self.progress_callback = progress_callback
        if not isinstance(context, Storage):
            context = Storage(context)
        self.context = context
        self.errors = []
    
    def error(self, row, col, message):
        message_seq = (self.sheet_name, row, col, message)
        self.errors.append(message_seq)
    
    def progress(self, *args):
        if self.progress_callback is not None:
            self.progress_callback(*args)
    
    def is_empty_row(self, sheet, row, num_cols=30):
        # we've picked 30 as an arbitrary number of columns to test so that
        # caller doesn't specify the number.
        for col in range(1, num_cols):
            try:
                if sheet.cell(row=row, column=col).value:
                    return False
            except (ValueError, IndexError):
                break
        return True
    
    def get_cell_value(self, sheet, row, col, default=no_data):
        try:
            return sheet.cell(row=row, column=col).value
        except ValueError:
            if default != no_data:
                return default
            raise
    
    def get_cell_and_found(self, sheet, row, col, default=''):
        try:
            value = sheet.cell(row=row, column=col).value
            return (value, True) if value else (default, False)
        except ValueError:
            return (default, False)
    
    def get_bool_from_cell(self, sheet, row, col):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not found:
            return None
        value = to_bool(value)
        if value is None:
            self.error(row, col, ERROR_NOT_BOOLEAN)
            return None
        return value
    
    def get_required_bool_from_cell(self, sheet, row, col):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not value:
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
            return None
        return self.get_bool_from_cell(sheet, row, col)
    
    def get_date_from_cell(self, sheet, row, col, default=None):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not found or value == '':
            return default
        
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        
        value = to_text(value)
        if value is not None:
            try:
                dt = [int(v) for v in value.replace('/', '-').split('-')]
                return date(*dt)
            except (TypeError, ValueError):
                self.error(row, col, ERROR_NOT_DATE)
        return None
    
    def get_required_date_from_cell(self, sheet, row, col):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not found or value == '':
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
            return None
        return self.get_date_from_cell(sheet, row, col)
    
    def get_enum_from_cell(self, sheet, row, col, enum_type, default=no_enum):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not found or value == '':
            if default is no_enum:
                self.error(row, col, ERROR_NOT_ENUM_OF(enum_type))
                return None
            return default
        value = to_enum(enum_type, value)
        if value is None:
            self.error(row, col, ERROR_NOT_ENUM_OF(enum_type))
        return value
    
    def get_required_enum_from_cell(self, sheet, row, col, enum_type):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not found or value == '':
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
            return None
        return self.get_enum_from_cell(sheet, row, col, enum_type)

    def get_float_found_valid(self, sheet, row, col, default=0.0):
        value, found = self.get_cell_and_found(sheet, row, col, default)
        valid = True
        if isinstance(value, str) and not value.strip():
            value, found = (default, False)
        if found:
            try:
                value = float(value)
            except (TypeError, ValueError):
                self.error(row, col, ERROR_NOT_FLOAT)
                valid, value = (False, default or None)
        return (value, found, valid)
    
    def get_float_from_cell(self, sheet, row, col, default=0.0):
        value, found, valid = self.get_float_found_valid(sheet, row, col, default)
        return value
    
    def get_required_float_from_cell(self, sheet, row, col):
        value, found, valid = self.get_float_found_valid(sheet, row, col)
        if not found:
            self.error(row, col, ERROR_NOT_FLOAT)
        return value
    
    def get_id_from_cell(self, sheet, row, col, default=''):
        value, found, valid = self.get_text_found_valid(sheet, row, col, default)
        return value.strip()
    
    def get_required_id_from_cell(self, sheet, row, col):
        value, found, valid = self.get_text_found_valid(sheet, row, col)
        if valid and not value:
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
        return value.strip()
    
    def get_ids_from_cell(self, sheet, row, col):
        value, found, valid = self.get_text_found_valid(sheet, row, col)
        if not value:
            return []
        return [p.strip() for p in value.split(',') if p.strip()]
    
    def get_required_ids_from_cell(self, sheet, row, col):
        value, found = self.get_cell_and_found(sheet, row, col)
        if not value:
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
            return []
        return self.get_ids_from_cell(sheet, row, col)
    
    def get_int_found_valid(self, sheet, row, col, default=0):
        value, found = self.get_cell_and_found(sheet, row, col, default)
        valid = True
        if isinstance(value, str) and not value.strip():
            value, found = (default, False)
        if found:
            try:
                value = int(value)
            except (TypeError, ValueError):
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    pass
            if isinstance(value, float):
                if int(value) != value:
                    self.error(row, col, ERROR_NOT_INT)
                    valid = False
                else:
                    value = int(value)
            elif not isinstance(value, int):
                self.error(row, col, ERROR_NOT_INT)
                valid, value = (False, None)
        return (value, found, valid)
    
    def get_int_from_cell(self, sheet, row, col, default=0):
        value, found, valid = self.get_int_found_valid(sheet, row, col, default)
        return value
    
    def get_required_int_from_cell(self, sheet, row, col):
        value, found, valid = self.get_int_found_valid(sheet, row, col)
        if not found:
            self.error(row, col, ERROR_NOT_INT)
        return value

    def get_text_found_valid(self, sheet, row, col, default=''):
        value, found = self.get_cell_and_found(sheet, row, col, default)
        valid = True
        if found:
            value = to_text(value)
            if not value:
                self.error(row, col, ERROR_NOT_UNICODE_OR_ASCII)
                valid, value = (False, default)
        return (value, found, valid)
    
    def get_text_from_cell(self, sheet, row, col, default=''):
        value, found, valid = self.get_text_found_valid(sheet, row, col, default)
        return value
    
    def get_required_text_from_cell(self, sheet, row, col):
        value, found, valid = self.get_text_found_valid(sheet, row, col)
        if valid and not value:
            self.error(row, col, ERROR_MISSING_REQUIRED_TEXT)
        return value

    def find_headers(self, sh, sample_headers, row_offset=1, max_row_check=30):
        norm_headers = [h.lower() for h in sample_headers]
        check_index, hdr_count = 1, len(norm_headers)

        for row in range(row_offset, sh.max_row + 1):
            if check_index > max_row_check:
                return -1
            cols = []
            for i in range(1, hdr_count + 1):
                value = sh.cell(row=row, column=i).value
                cols.append(str(value or '').lower())
            
            if norm_headers == cols:
                return row
            check_index += 1
        return -1

    def validate_column_headings(self, sh):
        headers = (self.headers or ('fake-@#$%^',))
        if not self.find_headers(sh, headers):
            raise ValidationError("Expected columns not found.")
    
    @property
    def sheet(self):
        if self.sheet_name and self.wb:
            if self.sheet_name in self.wb.sheetnames:
                return self.wb.get_sheet_by_name(self.sheet_name)
        return None

    def import_data(self, source):
        wb = source if isinstance(source, openpyxl.Workbook) else None
        if not wb and isinstance(source, str):
            if os.path.isfile(source):
                wb = openpyxl.load_workbook(source, read_only=True)
        
        if not wb:
            raise ValueError("`source` expected to be path to .xls(x) file " \
                             "or instance of openpyxl.Workbook.")
        self.wb = wb
        if self.sheet:
            self.process()


class IXCacheMixin(object):
    cache = {}
    def get_office_id(self, short_name):
        if short_name and short_name not in self.cache:
            office = BusinessOffice.objects.filter(short_name=short_name).first()
            if office is not None:
                self.cache[short_name] = office.id
        return self.cache.get(short_name, None)
    
    def get_state_id(self, state_code):
        if state_code not in self.cache:
            state = State.objects.filter(code=state_code).first()
            if state is not None:
                self.cache[state_code] = state.id
        return self.cache.get(state_code, None)
    
    def get_station_id(self, station_code):
        if station_code not in self.cache:
            station = Station.objects.filter(code=station_code).first()
            if station is not None:
                self.cache[station_code] = station.id
        return self.cache.get(station_code, None)
    
    def get_powerline_id(self, powerline_code):
        if powerline_code not in self.cache:
            powerline = Powerline.objects.filter(code=powerline_code).first()
            if powerline is not None:
                self.cache[powerline_code] = powerline.id
        return self.cache.get(powerline_code, None)
    
    def resolve_office(self, data, field):
        office_id, office_code = (None, data.pop(field))
        if office_code is not None:
            office_id = self.get_office_id(office_code)
        data[field] = office_id
    
    def resolve_state(self, data, field='addr_state_id'):
        state_code = data.pop(field)
        state_id = self.get_state_id(state_code)
        data[field] = state_id
    
    def resolve_station(self, data, field='source_station_id'):
        station_code = data.pop(field)
        if station_code is not None:
            station_id = self.get_station_id(station_code)
        data[field] = station_id
    
    def resolve_powerline(self, data, field='source_powerline_id'):
        powerline_code = data.pop(field)
        if powerline_code is not None:
            powerline_id = self.get_powerline_id(powerline_code)
        data[field] = powerline_id


class AccountIXHandler(IXlrdHandlerBase):
    sheet_name = 'accounts'
    table_name = 'enum_account'
    headers = ('sn', 'acct_no', 'book_code', 'cust_name', 'service_addr', 
               'cust_mobile', 'meter_no', 'tariff')

    def process(self):
        self.validate_column_headings()

        # read data into DataFrame
        try:
            import pandas as pd
            df = pd.read_excel(self._filepath, self.sheet_name)
            df.rename(columns={'mobile':'phone'}, inplace=True)
            df['date_created'] = datetime.today().strftime("%Y-%m-%d")
            del df['sn']
            
            conn_str = self.get_db_connstr()
            engine = create_engine(conn_str)
            df.to_sql(self.table_name, engine, if_exists='append', index=False)
        except Exception as ex:
            self.error(str(ex))
    
    def import_data(self, filepath):
        if not os.path.isfile(filepath):
            self.error("File not found: %s" % filepath)
            return 
        self._filepath = filepath
        self.process()


class UserIXHandler(IPyXlHandlerBase, IXCacheMixin):
    sheet_name = 'users'
    headers = ('sn', 'email', 'first_name', 'last_name', 'is_active', 'is_staff',
               'phone', 'location')
    
    def import_user(self, row, data):
        try:
            phone = data.pop('phone')
            location_id = self.get_office_id(data.pop('location'))
            profile = UserProfile(
                user=User.objects.create(**data),
                phone=phone, location_id=location_id
            )
            profile.save()
        except Exception as ex:
            message_fmt = "User could not be created. Err: %s"
            self.error(row, 0, message_fmt % str(ex))

    def process(self):
        sh = self.sheet
        header_index = self.find_headers(sh, self.headers)
        if not header_index:
            raise ValidationError("Expected columns not found.")
        
        with transaction.atomic():
            for row in range(header_index + 1, sh.max_row + 1):
                data = Storage({
                    'email': self.get_required_text_from_cell(sh, row, 2),
                    'first_name': self.get_required_text_from_cell(sh, row, 3),
                    'last_name': self.get_required_text_from_cell(sh, row, 4),
                    'is_active': self.get_required_bool_from_cell(sh, row, 5),
                    'is_staff': self.get_required_bool_from_cell(sh, row, 6),
                    'phone': self.get_text_from_cell(sh, row, 7),
                    'location': self.get_text_from_cell(sh, row, 8)
                })
                data.date_joined = datetime.today().isoformat(' ')[:19]
                data.username = data.email
                data.is_superuser = False
                data.password = '!t3mp!'
                self.import_user(row, data)


class OfficeIXHandler(IPyXlHandlerBase, IXCacheMixin):
    sheet_name = 'offices'
    headers = ('sn', 'code', 'short_name', 'name', 'level', 'parent', 'category',
               'email', 'phone', 'website')
        
    def import_office(self, row, data):
        try:
            # resolve addr_state
            self.resolve_state(data)
            self.resolve_office(data, 'parent_id')
            office = BusinessOffice(**data)
            office.save()
        except Exception as ex:
            message_fmt = "Office could not be created. Err: %s"
            self.error(row, 0, message_fmt % str(ex))

    def process(self):
        sh = self.sheet
        hdr_idx = self.find_headers(sh, self.headers)
        if not hdr_idx:
            raise ValidationError("Expected columns not found.")
        
        with transaction.atomic():
            for row in range(hdr_idx + 1, sh.max_row + 1):
                data = Storage({
                    'code': self.get_required_text_from_cell(sh, row, 2),
                    'short_name': self.get_required_text_from_cell(sh, row, 3),
                    'name': self.get_required_text_from_cell(sh, row, 4),
                    'level_id': self.get_required_text_from_cell(sh, row, 5),
                    'parent_id': self.get_id_from_cell(sh, row, 6) or None,
                    'category': self.get_text_from_cell(sh, row, 7),
                    'email': self.get_text_from_cell(sh, row, 8),
                    'phone': self.get_text_from_cell(sh, row, 9),
                    'website': self.get_text_from_cell(sh, row, 10),
                    'addr_street': self.get_text_from_cell(sh, row, 11),
                    'addr_town': self.get_required_text_from_cell(sh, row, 12),
                    'addr_state_id': self.get_required_id_from_cell(sh, row, 13),
                    'postal_code': self.get_text_from_cell(sh, row, 14),
                    'longitude': self.get_float_from_cell(sh, row, 15),
                    'latitude': self.get_float_from_cell(sh, row, 16),
                    'altitude': self.get_float_from_cell(sh, row, 17),
                    'gps_error': self.get_int_from_cell(sh, row, 18)
                })
                self.import_office(row, data)


class StationPowerlineIXHandler(IPyXlHandlerBase, IXCacheMixin):
    sheet_name = 'stations-lines'

    def _get_station_type(self, sh, row, column):
        value = self.get_required_text_from_cell(sh, row, column).title()
        for (code, name) in Station.STATION_CHOICES:
            if value == name:
                return code
        self.error(row, column, "Invalid station type encountered: %s" % value)
        return None
    
    def _get_line_type(self, sh, row, column):
        value = self.get_required_text_from_cell(sh, row, column).title()
        for (code, name) in Powerline.POWERLINE_CHOICES:
            if value == name:
                return code
        self.error(row, column, "Invalid powerline type encountered: %s" % value)
        return None
    
    def _get_voltage(self, sh, row, column):
        value = str(self.get_int_from_cell(sh, row, column) // 1000) + "KV"
        for (code, name) in Voltage.ALL_CHOICES:
            if value == name:
                return code
        self.error(row, column, "Invalid voltage value encountered: %s" % value)
        return None
    
    def _get_voltage_ratio(self, station_type):
        if station_type == Station.TRANSMISSION:
            return Voltage.Ratio.HVOLTL_MVOLTH
        elif station_type == Station.INJECTION:
            return Voltage.Ratio.MVOLTH_MVOLTL
        return Voltage.Ratio.MVOLTH_LVOLT
    
    def create_powerline(self, row, data):
        try:
            # a powerline can belong to multiple regions however in
            # CENTrak, they can only be assigned to one region
            field = 'region_id'
            data[field] = data[field][0] if data[field] else ''
            self.resolve_office(data, field)
            self.resolve_station(data)
            powerline = Powerline(**data)
            powerline.save()
        except Exception as ex:
            message_fmt = 'Powerline could not be created. Err: %s'
            self.error(row, 0, message_fmt % str(ex))
        
    def create_station(self, row, data):
        try:
            if data['type'] != Station.TRANSMISSION:
                self.resolve_powerline(data)
            self.resolve_office(data, 'region_id')
            self.resolve_state(data)
            station = Station(**data)
            station.save()
        except Exception as ex:
            message_fmt = 'Station could not be created. Err: %s'
            self.error(row, 0, message_fmt % str(ex))
    
    def import_powerline(self, sh, row):
        num_errors = len(self.errors)
        line_type = self._get_line_type(sh, row, 2)
        if num_errors < len(self.errors):
            return (row + 1)
        
        voltage = self._get_voltage(sh, row + 1, 2)
        if num_errors < len(self.errors):
            return (row + 1)
        
        row += 3
        num_errors = len(self.errors)
        while row <= sh.max_row:
            if self.is_empty_row(sh, row):
                break
            data = {'type': line_type, 'voltage': voltage}
            data['source_station_id'] = self.get_required_id_from_cell(sh, row, 1)
            data['code'] = self.get_required_id_from_cell(sh, row, 2)
            data['altcode'] = self.get_required_id_from_cell(sh, row, 3)
            data['region_id'] = self.get_ids_from_cell(sh, row, 4)
            data['name'] = self.get_required_text_from_cell(sh, row, 5)
            data['is_public'] = self.get_required_bool_from_cell(sh, row, 6)
            data['date_commissioned'] = self.get_date_from_cell(sh, row, 8)
            if num_errors < len(self.errors):
                break
            self.create_powerline(row, data)
            row += 1
        return row + 1

    def import_station(self, sh, row):
        num_errors = len(self.errors)
        station_type = self._get_station_type(sh, row, 2)
        voltage_ratio = self._get_voltage_ratio(station_type)
        if num_errors < len(self.errors):
            return row + 1
        row += 2
        num_errors = len(self.errors)
        while row <= sh.max_row:
            if self.is_empty_row(sh, row):
                break
            data = {'type': station_type, 'voltage_ratio': voltage_ratio}
            if station_type != Station.TRANSMISSION:
                data['source_powerline_id'] = self.get_required_id_from_cell(sh, row, 1)
            data['code'] = self.get_required_id_from_cell(sh, row, 2)
            data['altcode'] = self.get_required_id_from_cell(sh, row, 3)
            data['region_id'] = self.get_required_id_from_cell(sh, row, 4)
            data['name'] = self.get_required_text_from_cell(sh, row, 5)
            data['is_public'] = self.get_required_bool_from_cell(sh, row, 6)
            data['addr_state_id'] = self.get_required_text_from_cell(sh, row, 9)
            data['addr_street'] = self.get_text_from_cell(sh, row, 10)
            data['addr_town'] = self.get_text_from_cell(sh, row, 11)
            data['longitude'] = self.get_float_from_cell(sh, row, 13, None)
            data['latitude'] = self.get_float_from_cell(sh, row, 14, None)
            data['altitude'] = self.get_float_from_cell(sh, row, 15, None)
            data['gps_error'] = self.get_int_from_cell(sh, row, 16, None)
            data['date_commissioned'] = self.get_date_from_cell(sh, row, 17)
            if num_errors < len(self.errors):
                break
            self.create_station(row, data)
            row += 1
        return (row + 1)

    def process(self):
        sh, row, nrows = (self.sheet, 1, self.sheet.max_row)
        flags = ['lines', 'stations']   # provided in reverse bcos of pop
        with transaction.atomic():
            while row <= nrows:
                if not flags: break
                if self.is_empty_row(sh, row):
                    row += 1
                    continue
                
                value = sh.cell(row=row, column=1).value
                if value not in flags:
                    row += 1
                    continue
                
                while value != flags.pop():
                    pass    # keep popping ;-)
                
                if value == 'stations':
                    row = self.import_station(sh, row)
                elif value == 'lines':
                    row = self.import_powerline(sh, row)
                
                if self.errors:
                    transaction.rollback()
                    break
